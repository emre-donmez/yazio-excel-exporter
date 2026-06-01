from datetime import datetime, date
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


EXCEL_DATE_FORMAT = "dd.mm.yyyy"

MEAL_ORDER = {
    "breakfast": 0,
    "lunch": 1,
    "dinner": 2,
    "snack": 3
}


def _to_excel_date(value):
    """Convert ISO date string like 2026-05-18 to real Excel date."""
    if not value:
        return ""

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    if isinstance(value, str):
        try:
            return datetime.strptime(value[:10], "%Y-%m-%d").date()
        except ValueError:
            return value

    return value


def _date_sort_key(row: dict):
    """Return a safe date value for sorting."""
    value = _to_excel_date(row.get("date", ""))

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    return date.min


def _meal_sort_key(row: dict):
    """Return meal order for Details sheet sorting."""
    meal = row.get("meal", "")

    if not meal:
        return 99

    meal = str(meal).strip().lower()
    return MEAL_ORDER.get(meal, 99)


def _sort_summary_rows_by_date_desc(rows: list[dict]):
    """Sort Summary rows by date from newest to oldest."""
    return sorted(
        rows,
        key=_date_sort_key,
        reverse=True
    )


def _sort_detail_rows_by_date_desc_and_meal(rows: list[dict]):
    """
    Sort Details rows by:
    1. Date from newest to oldest
    2. Meal order inside the same day:
       Breakfast -> Lunch -> Dinner -> Snack
    """
    return sorted(
        rows,
        key=lambda row: (
            -_date_sort_key(row).toordinal(),
            _meal_sort_key(row)
        )
    )


def _sort_weight_rows_by_date_desc(rows: list[dict]):
    """Sort Weight Change rows by date from newest to oldest."""
    return sorted(
        rows,
        key=_date_sort_key,
        reverse=True
    )


def export_to_excel(
    summary_rows: list[dict],
    detail_rows: list[dict],
    output_path: str,
    weight_change_rows: Optional[list[dict]] = None
):
    """Export data to Excel file with Summary, Details, and Weight Change sheets."""
    wb = Workbook()

    summary_rows = _sort_summary_rows_by_date_desc(summary_rows)
    detail_rows = _sort_detail_rows_by_date_desc_and_meal(detail_rows)
    weight_change_rows = _sort_weight_rows_by_date_desc(weight_change_rows or [])

    _create_summary_sheet(wb, summary_rows)
    _create_detail_sheet(wb, detail_rows)
    _create_weight_change_sheet(wb, weight_change_rows)

    wb.save(output_path)
    print(f"Excel file saved to: {output_path}")


def _create_summary_sheet(wb: Workbook, rows: list[dict]):
    ws = wb.active
    ws.title = "Summary"

    headers = [
        "Date",
        "Calories (kcal)",
        "Protein (g)",
        "Carbs (g)",
        "Fat (g)",
        "Fiber (g)",
        "Calorie Goal",
        "Goal - Calories"
    ]

    fields = [
        "date",
        "calories",
        "protein",
        "carbs",
        "fat",
        "fiber",
        "calorie_goal",
        "goal_minus_calories"
    ]

    # Style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, field in enumerate(fields, 1):
            val = row_data.get(field, "")

            if field == "date":
                val = _to_excel_date(val)

            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

            if field == "date":
                cell.number_format = EXCEL_DATE_FORMAT
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.number_format = '#,##0.0'
                cell.alignment = Alignment(horizontal="right")

    # Alternating row colors
    light_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    for row_idx in range(2, len(rows) + 2):
        if row_idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = light_fill

    # Auto-fit column widths
    for col_idx in range(1, len(headers) + 1):
        max_len = len(headers[col_idx - 1])

        for row_idx in range(2, len(rows) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value

            if isinstance(val, date):
                val = val.strftime("%d.%m.%Y")

            val = str(val or "")
            max_len = max(max_len, len(val))

        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3

    # Freeze top row
    ws.freeze_panes = "A2"


def _create_detail_sheet(wb: Workbook, rows: list[dict]):
    ws = wb.create_sheet("Details")

    headers = [
        "Date",
        "Meal",
        "Food Name",
        "Producer",
        "Amount (g)",
        "Calories (kcal)",
        "Protein (g)",
        "Carbs (g)",
        "Fat (g)",
        "Fiber (g)",
        "AI Generated"
    ]

    fields = [
        "date",
        "meal",
        "food_name",
        "producer",
        "amount",
        "calories",
        "protein",
        "carbs",
        "fat",
        "fiber",
        "ai_generated"
    ]

    # Style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, field in enumerate(fields, 1):
            val = row_data.get(field, "")

            if field == "date":
                val = _to_excel_date(val)

            if field == "ai_generated":
                val = "Yes" if val else "No"

            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

            if field == "date":
                cell.number_format = EXCEL_DATE_FORMAT
                cell.alignment = Alignment(horizontal="center")

            elif field in ("calories", "protein", "carbs", "fat", "fiber", "amount"):
                cell.number_format = '#,##0.0'
                cell.alignment = Alignment(horizontal="right")

            elif field == "ai_generated":
                cell.alignment = Alignment(horizontal="center")

                if val == "Yes":
                    cell.font = Font(color="FF0000")

            else:
                cell.alignment = Alignment(horizontal="left")

    # Alternating row colors
    light_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    for row_idx in range(2, len(rows) + 2):
        if row_idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = light_fill

    # Auto-fit column widths
    for col_idx in range(1, len(headers) + 1):
        max_len = len(headers[col_idx - 1])

        for row_idx in range(2, min(len(rows) + 2, 100)):
            val = ws.cell(row=row_idx, column=col_idx).value

            if isinstance(val, date):
                val = val.strftime("%d.%m.%Y")

            val = str(val or "")
            max_len = max(max_len, len(val))

        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 50)

    # Freeze top row
    ws.freeze_panes = "A2"


def _create_weight_change_sheet(wb: Workbook, rows: list[dict]):
    ws = wb.create_sheet("Weight Change")

    headers = [
        "Date",
        "Weight (kg)",
        "Change (kg)",
        "Total Change (kg)"
    ]

    fields = [
        "date",
        "weight",
        "change",
        "total_change"
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, field in enumerate(fields, 1):
            val = row_data.get(field, "")

            if field == "date":
                val = _to_excel_date(val)

            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

            if field == "date":
                cell.number_format = EXCEL_DATE_FORMAT
                cell.alignment = Alignment(horizontal="center")
            elif field == "weight":
                cell.number_format = '#,##0.0'
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.number_format = '+#,##0.0;-#,##0.0;0.0'
                cell.alignment = Alignment(horizontal="right")

    light_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    for row_idx in range(2, len(rows) + 2):
        if row_idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = light_fill

    for col_idx in range(1, len(headers) + 1):
        max_len = len(headers[col_idx - 1])

        for row_idx in range(2, len(rows) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value

            if isinstance(val, date):
                val = val.strftime("%d.%m.%Y")

            val = str(val or "")
            max_len = max(max_len, len(val))

        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3

    ws.freeze_panes = "A2"
